#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""Excel exporter with advanced formatting and charts."""

from __future__ import annotations


import logging
from pathlib import Path
from collections import Counter
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference

from .base import BaseExporter
from ..models import ProjectData, AccessPoint, Antenna
from ..analytics import (
    GroupingAnalytics,
    CoverageAnalytics,
    MountingAnalytics,
    RadioAnalytics,
)

logger = logging.getLogger(__name__)


class ExcelExporter(BaseExporter):
    """Export project data to Excel files with formatting and charts.

    Creates Excel workbook with multiple sheets:
    - Summary: Project overview and statistics
    - Access Points: Detailed AP list with tags
    - Antennas: Antenna list
    - By Floor: Grouped by floor with chart
    - By Color: Grouped by color with chart
    - By Vendor: Grouped by vendor with chart
    - By Model: Grouped by model with chart
    """

    # Styles for headers
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    # Styles for summary section headers
    SECTION_FONT = Font(bold=True, size=12, color="366092")

    # Borders
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    @property
    def format_name(self) -> str:
        """Human-readable name of the export format."""
        return "Excel"

    def export(self, project_data: ProjectData) -> list[Path]:
        """Export project data to Excel file.

        Args:
            project_data: Processed project data to export

        Returns:
            List with single Excel file path

        Raises:
            IOError: If file writing fails
        """
        output_file = self._get_output_filename(
            project_data.project_name, f"{project_data.project_name}.xlsx"
        )

        logger.info(f"Creating Excel file: {output_file}")

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create sheets in order
        self._create_summary_sheet(wb, project_data)
        self._create_access_points_sheet(
            wb, project_data.access_points, project_data.project_name
        )
        self._create_detailed_access_points_sheet(
            wb, project_data.access_points, project_data.project_name
        )
        self._create_antennas_sheet(
            wb, project_data.antennas, project_data.project_name
        )
        self._create_grouped_sheets(wb, project_data.access_points)

        # Create analytics sheet if data available
        has_mounting_data = any(
            ap.mounting_height is not None for ap in project_data.access_points
        )
        has_radio_data = len(project_data.radios) > 0
        if has_mounting_data or has_radio_data:
            self._create_analytics_sheet(
                wb, project_data.access_points, project_data.radios
            )

        # Save workbook
        try:
            wb.save(output_file)
            logger.info(f"Excel file created successfully: {output_file}")
        except Exception as e:
            logger.error(f"Failed to save Excel file: {e}")
            raise IOError(f"Failed to save Excel file: {e}")

        self.log_export_success([output_file])

        return [output_file]

    def _apply_header_style(self, ws, row: int = 1):
        """Apply header style to specified row.

        Args:
            ws: Worksheet
            row: Row number (1-indexed)
        """
        for cell in ws[row]:
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

    def _auto_size_columns(self, ws):
        """Auto-size all columns based on content.

        Args:
            ws: Worksheet
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Set width with min 10, max 50 characters
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _apply_autofilter(self, ws):
        """Apply autofilter to the sheet.

        Args:
            ws: Worksheet
        """
        if ws.max_row > 1:  # Only if there's data
            ws.auto_filter.ref = ws.dimensions

    def _freeze_header(self, ws, row: int = 2):
        """Freeze header row.

        Args:
            ws: Worksheet
            row: Row number to freeze (all rows above will be frozen)
        """
        ws.freeze_panes = ws[f"A{row}"]

    def _apply_borders(
        self,
        ws,
        min_row: int = 1,
        max_row: int = None,
        min_col: int = 1,
        max_col: int = None,
    ):
        """Apply borders to range of cells.

        Args:
            ws: Worksheet
            min_row: Starting row (1-indexed)
            max_row: Ending row (None for all)
            min_col: Starting column (1-indexed)
            max_col: Ending column (None for all)
        """
        if max_row is None:
            max_row = ws.max_row
        if max_col is None:
            max_col = ws.max_column

        for row in ws.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            for cell in row:
                cell.border = self.THIN_BORDER

    def _create_summary_sheet(self, wb: Workbook, project_data: ProjectData):
        """Create summary sheet with project statistics.

        Args:
            wb: Workbook
            project_data: Project data
        """
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws["A1"] = "Project Summary"
        ws["A1"].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells("A1:B1")

        row = 3

        # Project metadata section (if available)
        if project_data.metadata:
            ws[f"A{row}"] = "Project Information"
            ws[f"A{row}"].font = self.SECTION_FONT
            row += 1

            if project_data.metadata.name:
                ws[f"A{row}"] = "Project Name:"
                ws[f"B{row}"] = project_data.metadata.name
                ws[f"A{row}"].font = Font(bold=True)
                row += 1

            if project_data.metadata.customer:
                ws[f"A{row}"] = "Customer:"
                ws[f"B{row}"] = project_data.metadata.customer
                ws[f"A{row}"].font = Font(bold=True)
                row += 1

            if project_data.metadata.location:
                ws[f"A{row}"] = "Location:"
                ws[f"B{row}"] = project_data.metadata.location
                ws[f"A{row}"].font = Font(bold=True)
                row += 1

            if project_data.metadata.responsible_person:
                ws[f"A{row}"] = "Responsible Person:"
                ws[f"B{row}"] = project_data.metadata.responsible_person
                ws[f"A{row}"].font = Font(bold=True)
                row += 1

            if project_data.metadata.schema_version:
                ws[f"A{row}"] = "Schema Version:"
                ws[f"B{row}"] = project_data.metadata.schema_version
                ws[f"A{row}"].font = Font(bold=True)
                row += 1

            row += 1  # Add spacing

        # Project statistics section
        ws[f"A{row}"] = "Project Statistics"
        ws[f"A{row}"].font = self.SECTION_FONT
        row += 1

        ws[f"A{row}"] = "File Name:"
        ws[f"B{row}"] = project_data.project_name
        ws[f"A{row}"].font = Font(bold=True)

        row += 1
        ws[f"A{row}"] = "Total Access Points:"
        ws[f"B{row}"] = len(project_data.access_points)
        ws[f"A{row}"].font = Font(bold=True)

        row += 1
        ws[f"A{row}"] = "Total Antennas:"
        ws[f"B{row}"] = len(project_data.antennas)
        ws[f"A{row}"].font = Font(bold=True)

        row += 1
        ws[f"A{row}"] = "Unique AP Models:"
        unique_models = len(set(ap.model for ap in project_data.access_points))
        ws[f"B{row}"] = unique_models
        ws[f"A{row}"].font = Font(bold=True)

        # Statistics by vendor
        row += 2
        ws[f"A{row}"] = "Distribution by Vendor"
        ws[f"A{row}"].font = self.SECTION_FONT

        row += 1
        vendor_counts = Counter(ap.vendor for ap in project_data.access_points)
        ws[f"A{row}"] = "Vendor"
        ws[f"B{row}"] = "Count"
        ws[f"C{row}"] = "Percentage"
        self._apply_header_style(ws, row)

        total_aps = len(project_data.access_points)
        for vendor, count in sorted(
            vendor_counts.items(), key=lambda x: x[1], reverse=True
        ):
            row += 1
            percentage = (count / total_aps * 100) if total_aps > 0 else 0
            ws[f"A{row}"] = vendor
            ws[f"B{row}"] = count
            ws[f"C{row}"] = f"{percentage:.1f}%"

        # Statistics by floor
        row += 2
        ws[f"A{row}"] = "Distribution by Floor"
        ws[f"A{row}"].font = self.SECTION_FONT

        row += 1
        floor_counts = Counter(ap.floor_name for ap in project_data.access_points)
        ws[f"A{row}"] = "Floor"
        ws[f"B{row}"] = "Count"
        ws[f"C{row}"] = "Percentage"
        self._apply_header_style(ws, row)

        for floor, count in sorted(
            floor_counts.items(), key=lambda x: x[1], reverse=True
        ):
            row += 1
            percentage = (count / total_aps * 100) if total_aps > 0 else 0
            ws[f"A{row}"] = floor
            ws[f"B{row}"] = count
            ws[f"C{row}"] = f"{percentage:.1f}%"

        # Auto-size columns
        self._auto_size_columns(ws)

    def _create_access_points_sheet(
        self, wb: Workbook, access_points: list[AccessPoint], project_name: str
    ):
        """Create detailed access points sheet.

        Args:
            wb: Workbook
            access_points: List of access points
            project_name: Project name
        """
        ws = wb.create_sheet("Access Points")

        # Headers
        headers = ["Vendor", "Model", "Floor", "Color", "Tags", "Quantity"]
        ws.append(headers)
        self._apply_header_style(ws)

        # Group and count APs
        ap_tuples = [
            (
                ap.vendor,
                ap.model,
                ap.floor_name,
                ap.color,
                frozenset(str(tag) for tag in ap.tags),
            )
            for ap in access_points
        ]
        ap_counts = Counter(ap_tuples)

        logger.info(
            f"Exporting {len(access_points)} access points ({len(ap_counts)} unique) to Excel"
        )

        # Write data
        for (vendor, model, floor, color, tags), count in sorted(ap_counts.items()):
            tags_str = "; ".join(sorted(tags)) if tags else ""
            ws.append([vendor, model, floor, color or "", tags_str, count])

        # Apply formatting
        self._auto_size_columns(ws)
        self._apply_autofilter(ws)
        self._freeze_header(ws)
        self._apply_borders(ws)

    def _create_detailed_access_points_sheet(
        self, wb: Workbook, access_points: list[AccessPoint], project_name: str
    ):
        """Create detailed access points sheet with installation parameters.

        Each row represents a single access point with all its properties including
        mounting height, azimuth, tilt, and location coordinates.

        Args:
            wb: Workbook
            access_points: List of access points
            project_name: Project name
        """
        ws = wb.create_sheet("AP Installation Details")

        # Headers with all fields
        headers = [
            "AP Name",
            "Vendor",
            "Model",
            "Floor",
            "Location X (m)",
            "Location Y (m)",
            "Mounting Height (m)",
            "Azimuth (°)",
            "Tilt (°)",
            "Color",
            "Tags",
            "Enabled",
        ]
        ws.append(headers)
        self._apply_header_style(ws)

        logger.info(
            f"Exporting {len(access_points)} access points with detailed installation parameters"
        )

        # Write data - one row per AP
        for ap in access_points:
            # Format tags as "Key1:Value1; Key2:Value2"
            tags_str = (
                "; ".join(str(tag) for tag in sorted(ap.tags, key=lambda t: t.key))
                if ap.tags
                else ""
            )

            # Prepare numeric values (will be None if not set)
            location_x = ap.location_x if ap.location_x is not None else None
            location_y = ap.location_y if ap.location_y is not None else None
            mounting_height = (
                ap.mounting_height if ap.mounting_height is not None else None
            )
            azimuth = ap.azimuth if ap.azimuth is not None else None
            tilt = ap.tilt if ap.tilt is not None else None

            ws.append(
                [
                    ap.name or "",
                    ap.vendor,
                    ap.model,
                    ap.floor_name,
                    location_x,
                    location_y,
                    mounting_height,
                    azimuth,
                    tilt,
                    ap.color or "",
                    tags_str,
                    "Yes" if ap.enabled else "No",
                ]
            )

        # Apply formatting
        self._auto_size_columns(ws)
        self._apply_autofilter(ws)
        self._freeze_header(ws)
        self._apply_borders(ws)

        # Apply number formatting to numeric columns
        for row in range(2, ws.max_row + 1):  # Skip header row
            # Location X, Y (columns 5, 6) - 2 decimal places
            if ws.cell(row, 5).value is not None:
                ws.cell(row, 5).number_format = "0.00"
            if ws.cell(row, 6).value is not None:
                ws.cell(row, 6).number_format = "0.00"
            # Mounting Height (column 7) - 2 decimal places
            if ws.cell(row, 7).value is not None:
                ws.cell(row, 7).number_format = "0.00"
            # Azimuth, Tilt (columns 8, 9) - 1 decimal place
            if ws.cell(row, 8).value is not None:
                ws.cell(row, 8).number_format = "0.0"
            if ws.cell(row, 9).value is not None:
                ws.cell(row, 9).number_format = "0.0"

        logger.info("Detailed AP installation sheet created")

    def _create_antennas_sheet(
        self, wb: Workbook, antennas: list[Antenna], project_name: str
    ):
        """Create antennas sheet.

        Args:
            wb: Workbook
            antennas: List of antennas
            project_name: Project name
        """
        ws = wb.create_sheet("Antennas")

        # Headers
        headers = ["Antenna Model", "Quantity"]
        ws.append(headers)
        self._apply_header_style(ws)

        # Count antennas
        antenna_names = [antenna.name for antenna in antennas]
        antenna_counts = Counter(antenna_names)

        logger.info(
            f"Exporting {len(antennas)} antennas ({len(antenna_counts)} unique) to Excel"
        )

        # Write data
        for antenna_name, count in sorted(antenna_counts.items()):
            ws.append([antenna_name, count])

        # Apply formatting
        self._auto_size_columns(ws)
        self._apply_autofilter(ws)
        self._freeze_header(ws)
        self._apply_borders(ws)

    def _create_grouped_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        grouped_data: dict[str, int],
        dimension_name: str,
        add_chart: bool = True,
    ):
        """Create a grouped sheet with counts, percentages, and optional chart.

        Args:
            wb: Workbook
            sheet_name: Name of the sheet
            grouped_data: Dictionary with counts by dimension
            dimension_name: Name of the dimension (e.g., "Floor", "Vendor")
            add_chart: Whether to add a chart
        """
        ws = wb.create_sheet(sheet_name)

        # Headers
        headers = [dimension_name, "Count", "Percentage"]
        ws.append(headers)
        self._apply_header_style(ws)

        # Calculate total
        total = sum(grouped_data.values())

        # Write data sorted by count (descending)
        for key, count in sorted(
            grouped_data.items(), key=lambda x: x[1], reverse=True
        ):
            percentage = (count / total * 100) if total > 0 else 0
            ws.append([str(key), count, f"{percentage:.1f}%"])

        # Apply formatting
        self._auto_size_columns(ws)
        self._apply_autofilter(ws)
        self._freeze_header(ws)
        self._apply_borders(ws)

        # Add chart if requested and there's data
        if add_chart and len(grouped_data) > 0:
            self._add_chart(ws, dimension_name)

    def _add_chart(self, ws, dimension_name: str):
        """Add appropriate chart to worksheet.

        Args:
            ws: Worksheet
            dimension_name: Name of the dimension
        """
        # Use pie chart for vendors, bar chart for others
        if dimension_name == "Vendor":
            chart = PieChart()
            chart.title = f"Distribution by {dimension_name}"
            chart.style = 10
            chart.height = 10
            chart.width = 15
        else:
            chart = BarChart()
            chart.type = "col"  # Column chart
            chart.title = f"Count by {dimension_name}"
            chart.style = 10
            chart.height = 10
            chart.width = 15
            chart.y_axis.title = "Count"
            chart.x_axis.title = dimension_name

        # Data for chart
        labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)

        # Position chart to the right of data
        ws.add_chart(chart, "E2")

    def _create_grouped_sheets(self, wb: Workbook, access_points: list[AccessPoint]):
        """Create all grouped sheets.

        Args:
            wb: Workbook
            access_points: List of access points
        """
        analytics = GroupingAnalytics()

        # By Floor
        floor_data = analytics.group_by_floor(access_points)
        self._create_grouped_sheet(wb, "By Floor", floor_data, "Floor")

        # By Color
        color_data = analytics.group_by_color(access_points)
        self._create_grouped_sheet(wb, "By Color", color_data, "Color")

        # By Vendor
        vendor_data = analytics.group_by_vendor(access_points)
        self._create_grouped_sheet(wb, "By Vendor", vendor_data, "Vendor")

        # By Model
        model_data = analytics.group_by_model(access_points)
        self._create_grouped_sheet(wb, "By Model", model_data, "Model")

        logger.info("Created 4 grouped sheets with charts")

    def _create_analytics_sheet(
        self, wb: Workbook, access_points: list[AccessPoint], radios: list
    ):
        """Create analytics sheet with mounting, coverage and radio metrics.

        Args:
            wb: Workbook to add sheet to
            access_points: List of access points
            radios: List of radios
        """
        ws = wb.create_sheet("Analytics")
        logger.info("Creating Analytics sheet")

        # Calculate mounting metrics
        mounting_metrics = MountingAnalytics.calculate_mounting_metrics(access_points)
        height_distribution = MountingAnalytics.group_by_height_range(access_points)
        installation_summary = MountingAnalytics.get_installation_summary(access_points)

        # Calculate radio metrics
        radio_metrics = None
        if radios:
            radio_metrics = RadioAnalytics.calculate_radio_metrics(radios)

        row = 1

        # Title
        ws.cell(row, 1, "INSTALLATION & MOUNTING ANALYTICS")
        ws.cell(row, 1).font = Font(bold=True, size=14, color="366092")
        row += 2

        # Mounting Metrics Section
        ws.cell(row, 1, "Mounting Metrics")
        ws.cell(row, 1).font = self.SECTION_FONT
        row += 1

        ws.cell(row, 1, "Metric")
        ws.cell(row, 2, "Value")
        ws.cell(row, 3, "Unit")
        self._apply_header_style(ws, row)
        row += 1

        if mounting_metrics.avg_height is not None:
            ws.cell(row, 1, "Average Mounting Height")
            ws.cell(row, 2, round(mounting_metrics.avg_height, 2))
            ws.cell(row, 3, "meters")
            row += 1

            ws.cell(row, 1, "Minimum Height")
            ws.cell(row, 2, round(mounting_metrics.min_height, 2))
            ws.cell(row, 3, "meters")
            row += 1

            ws.cell(row, 1, "Maximum Height")
            ws.cell(row, 2, round(mounting_metrics.max_height, 2))
            ws.cell(row, 3, "meters")
            row += 1

            ws.cell(row, 1, "Height Variance")
            ws.cell(row, 2, round(mounting_metrics.height_variance, 4))
            ws.cell(row, 3, "m²")
            row += 1

        ws.cell(row, 1, "APs with Height Data")
        ws.cell(row, 2, mounting_metrics.aps_with_height)
        ws.cell(row, 3, "count")
        row += 1

        if mounting_metrics.avg_azimuth is not None:
            ws.cell(row, 1, "Average Azimuth")
            ws.cell(row, 2, round(mounting_metrics.avg_azimuth, 1))
            ws.cell(row, 3, "degrees")
            row += 1

        if mounting_metrics.avg_tilt is not None:
            ws.cell(row, 1, "Average Tilt")
            ws.cell(row, 2, round(mounting_metrics.avg_tilt, 1))
            ws.cell(row, 3, "degrees")
            row += 1

        # Apply borders to metrics table
        self._apply_borders(ws, row - 7, row - 1, 1, 3)
        row += 1

        # Height Distribution Section
        ws.cell(row, 1, "Height Distribution")
        ws.cell(row, 1).font = self.SECTION_FONT
        row += 1

        ws.cell(row, 1, "Height Range")
        ws.cell(row, 2, "AP Count")
        self._apply_header_style(ws, row)
        dist_start_row = row + 1
        row += 1

        for range_name in [
            "< 2.5m",
            "2.5-3.5m",
            "3.5-4.5m",
            "4.5-6.0m",
            "> 6.0m",
            "Unknown",
        ]:
            count = height_distribution.get(range_name, 0)
            if count > 0:
                ws.cell(row, 1, range_name)
                ws.cell(row, 2, count)
                row += 1

        dist_end_row = row - 1

        # Apply borders to distribution table
        self._apply_borders(ws, dist_start_row - 1, dist_end_row, 1, 2)

        # Add chart for height distribution
        if dist_end_row >= dist_start_row:
            chart = BarChart()
            chart.title = "Height Distribution"
            chart.y_axis.title = "Number of APs"
            chart.x_axis.title = "Height Range"

            # Data references
            data = Reference(
                ws, min_col=2, min_row=dist_start_row - 1, max_row=dist_end_row
            )
            categories = Reference(
                ws, min_col=1, min_row=dist_start_row, max_row=dist_end_row
            )

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 10
            chart.width = 20

            ws.add_chart(chart, f"E{dist_start_row}")

        row += 1

        # Installation Summary Section
        ws.cell(row, 1, "Installation Summary")
        ws.cell(row, 1).font = self.SECTION_FONT
        row += 1

        ws.cell(row, 1, "Metric")
        ws.cell(row, 2, "Value")
        self._apply_header_style(ws, row)
        row += 1

        summary_start_row = row
        ws.cell(row, 1, "Total APs")
        ws.cell(row, 2, installation_summary["total_aps"])
        row += 1

        ws.cell(row, 1, "APs with Tilt Data")
        ws.cell(row, 2, installation_summary["aps_with_tilt"])
        row += 1

        ws.cell(row, 1, "APs with Azimuth Data")
        ws.cell(row, 2, installation_summary["aps_with_azimuth"])
        row += 1

        ws.cell(row, 1, "APs Requiring Height Adjustment")
        ws.cell(row, 2, installation_summary["aps_requiring_height_adjustment"])
        # Highlight if there are APs requiring adjustment
        if installation_summary["aps_requiring_height_adjustment"] > 0:
            ws.cell(row, 2).fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )
        row += 1

        # Apply borders to summary table
        self._apply_borders(ws, summary_start_row - 1, row - 1, 1, 2)

        # Radio Analytics Section
        if radio_metrics:
            row += 2

            # Radio Configuration Title
            ws.cell(row, 1, "RADIO CONFIGURATION ANALYTICS")
            ws.cell(row, 1).font = Font(bold=True, size=14, color="366092")
            row += 2

            # Radio Metrics Summary
            ws.cell(row, 1, "Radio Metrics")
            ws.cell(row, 1).font = self.SECTION_FONT
            row += 1

            ws.cell(row, 1, "Metric")
            ws.cell(row, 2, "Value")
            ws.cell(row, 3, "Unit")
            self._apply_header_style(ws, row)
            row += 1

            radio_summary_start = row
            ws.cell(row, 1, "Total Radios")
            ws.cell(row, 2, radio_metrics.total_radios)
            ws.cell(row, 3, "count")
            row += 1

            if radio_metrics.avg_tx_power:
                ws.cell(row, 1, "Average TX Power")
                ws.cell(row, 2, round(radio_metrics.avg_tx_power, 1))
                ws.cell(row, 3, "dBm")
                row += 1

                ws.cell(row, 1, "Min TX Power")
                ws.cell(row, 2, round(radio_metrics.min_tx_power, 1))
                ws.cell(row, 3, "dBm")
                row += 1

                ws.cell(row, 1, "Max TX Power")
                ws.cell(row, 2, round(radio_metrics.max_tx_power, 1))
                ws.cell(row, 3, "dBm")
                row += 1

            self._apply_borders(ws, radio_summary_start - 1, row - 1, 1, 3)
            row += 1

            # Frequency Bands Section
            if radio_metrics.band_distribution:
                ws.cell(row, 1, "Frequency Bands")
                ws.cell(row, 1).font = self.SECTION_FONT
                row += 1

                ws.cell(row, 1, "Band")
                ws.cell(row, 2, "Count")
                ws.cell(row, 3, "Percentage")
                self._apply_header_style(ws, row)
                band_start_row = row + 1
                row += 1

                for band, count in sorted(radio_metrics.band_distribution.items()):
                    percentage = (
                        (count / radio_metrics.total_radios * 100)
                        if radio_metrics.total_radios > 0
                        else 0
                    )
                    ws.cell(row, 1, band)
                    ws.cell(row, 2, count)
                    ws.cell(row, 3, f"{percentage:.1f}%")
                    row += 1

                band_end_row = row - 1
                self._apply_borders(ws, band_start_row - 1, band_end_row, 1, 3)

                # Add Pie Chart for frequency bands
                if band_end_row >= band_start_row:
                    chart = PieChart()
                    chart.title = "Frequency Band Distribution"
                    data = Reference(
                        ws, min_col=2, min_row=band_start_row - 1, max_row=band_end_row
                    )
                    labels = Reference(
                        ws, min_col=1, min_row=band_start_row, max_row=band_end_row
                    )
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(labels)
                    chart.height = 10
                    chart.width = 15
                    ws.add_chart(chart, f"E{band_start_row}")

                row += 1

            # Wi-Fi Standards Section
            if radio_metrics.standard_distribution:
                ws.cell(row, 1, "Wi-Fi Standards")
                ws.cell(row, 1).font = self.SECTION_FONT
                row += 1

                ws.cell(row, 1, "Standard")
                ws.cell(row, 2, "Count")
                ws.cell(row, 3, "Percentage")
                self._apply_header_style(ws, row)
                standard_start_row = row + 1
                row += 1

                for standard, count in sorted(
                    radio_metrics.standard_distribution.items()
                ):
                    percentage = (
                        (count / radio_metrics.total_radios * 100)
                        if radio_metrics.total_radios > 0
                        else 0
                    )
                    ws.cell(row, 1, standard)
                    ws.cell(row, 2, count)
                    ws.cell(row, 3, f"{percentage:.1f}%")
                    row += 1

                standard_end_row = row - 1
                self._apply_borders(ws, standard_start_row - 1, standard_end_row, 1, 3)

                # Add Bar Chart for Wi-Fi standards
                if standard_end_row >= standard_start_row:
                    chart = BarChart()
                    chart.title = "Wi-Fi Standards Distribution"
                    chart.y_axis.title = "Number of Radios"
                    data = Reference(
                        ws,
                        min_col=2,
                        min_row=standard_start_row - 1,
                        max_row=standard_end_row,
                    )
                    categories = Reference(
                        ws,
                        min_col=1,
                        min_row=standard_start_row,
                        max_row=standard_end_row,
                    )
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(categories)
                    chart.height = 10
                    chart.width = 15
                    ws.add_chart(chart, f"E{standard_start_row + 12}")

                row += 1

            # Channel Widths Section
            if radio_metrics.channel_width_distribution:
                ws.cell(row, 1, "Channel Widths")
                ws.cell(row, 1).font = self.SECTION_FONT
                row += 1

                ws.cell(row, 1, "Width (MHz)")
                ws.cell(row, 2, "Count")
                self._apply_header_style(ws, row)
                width_start_row = row + 1
                row += 1

                for width, count in sorted(
                    radio_metrics.channel_width_distribution.items()
                ):
                    ws.cell(row, 1, f"{width} MHz")
                    ws.cell(row, 2, count)
                    row += 1

                width_end_row = row - 1
                self._apply_borders(ws, width_start_row - 1, width_end_row, 1, 2)

        # Auto-size columns
        self._auto_size_columns(ws)

        logger.info(
            "Analytics sheet created with mounting and radio metrics and charts"
        )
