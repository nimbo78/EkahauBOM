"""Exporters for comparison results."""

from __future__ import annotations

import base64
import csv
import json
import logging
from pathlib import Path
from typing import Optional

from ekahau_bom.comparison.models import APChange, ChangeStatus, ComparisonResult

logger = logging.getLogger(__name__)


class ComparisonExporter:
    """Base class for comparison exporters."""

    def __init__(self, output_dir: Path, project_name: str = "comparison"):
        """Initialize exporter.

        Args:
            output_dir: Directory for output files
            project_name: Base name for output files
        """
        self.output_dir = output_dir
        self.project_name = self._sanitize_filename(project_name)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _sanitize_filename(self, name: str) -> str:
        """Remove invalid characters from filename."""
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            name = name.replace(char, "_")
        return name


class CSVComparisonExporter(ComparisonExporter):
    """Export comparison results to CSV format."""

    def export(self, comparison: ComparisonResult) -> Path:
        """Export comparison to CSV file.

        Args:
            comparison: ComparisonResult to export

        Returns:
            Path to generated CSV file
        """
        output_path = self.output_dir / f"{self.project_name}_comparison.csv"

        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            # Header
            writer.writerow(["AP Name", "Status", "Floor", "Details"])

            # Sort changes: removed first, then added, modified, moved, renamed
            status_order = {
                ChangeStatus.REMOVED: 0,
                ChangeStatus.ADDED: 1,
                ChangeStatus.MODIFIED: 2,
                ChangeStatus.MOVED: 3,
                ChangeStatus.RENAMED: 4,
                ChangeStatus.UNCHANGED: 5,
            }
            sorted_changes = sorted(
                comparison.ap_changes,
                key=lambda c: (status_order.get(c.status, 99), c.floor_name, c.ap_name),
            )

            # Write changes (exclude unchanged)
            for change in sorted_changes:
                if change.status == ChangeStatus.UNCHANGED:
                    continue

                writer.writerow(
                    [
                        change.ap_name,
                        change.status.value,
                        change.floor_name,
                        change.get_details_string(),
                    ]
                )

        logger.info(f"Exported comparison to CSV: {output_path}")
        return output_path


class ExcelComparisonExporter(ComparisonExporter):
    """Export comparison results to Excel format."""

    def export(self, comparison: ComparisonResult, diff_images: list[Path] = None) -> Path:
        """Export comparison to Excel file.

        Args:
            comparison: ComparisonResult to export
            diff_images: Optional list of diff visualization paths

        Returns:
            Path to generated Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl required for Excel export: pip install openpyxl")

        output_path = self.output_dir / f"{self.project_name}_comparison.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Status colors
        status_fills = {
            ChangeStatus.ADDED: PatternFill(start_color="C6EFCE", fill_type="solid"),  # Green
            ChangeStatus.REMOVED: PatternFill(start_color="FFC7CE", fill_type="solid"),  # Red
            ChangeStatus.MODIFIED: PatternFill(start_color="FFEB9C", fill_type="solid"),  # Yellow
            ChangeStatus.MOVED: PatternFill(start_color="D9E1F2", fill_type="solid"),  # Blue
            ChangeStatus.RENAMED: PatternFill(start_color="FFD699", fill_type="solid"),  # Orange
        }

        # Summary sheet
        summary_data = [
            ["Comparison Summary"],
            [],
            ["Old Project", comparison.project_a_name],
            ["New Project", comparison.project_b_name],
            ["Comparison Date", comparison.comparison_timestamp.strftime("%Y-%m-%d %H:%M")],
            [],
            ["Change Statistics"],
            ["Total APs (Old)", comparison.inventory_change.old_total_aps],
            ["Total APs (New)", comparison.inventory_change.new_total_aps],
            ["Difference", comparison.inventory_change.ap_count_diff],
            [],
            ["Added", comparison.inventory_change.aps_added],
            ["Removed", comparison.inventory_change.aps_removed],
            ["Modified", comparison.inventory_change.aps_modified],
            ["Moved", comparison.inventory_change.aps_moved],
            ["Renamed", comparison.inventory_change.aps_renamed],
            ["Unchanged", comparison.inventory_change.aps_unchanged],
        ]

        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx in [1, 7]:  # Headers
                    cell.font = Font(bold=True, size=14)

        # Adjust column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30

        # Changes sheet
        ws_changes = wb.create_sheet("Changes")
        headers = ["AP Name", "Status", "Floor", "Details"]

        for col_idx, header in enumerate(headers, 1):
            cell = ws_changes.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Sort and write changes
        status_order = {
            ChangeStatus.REMOVED: 0,
            ChangeStatus.ADDED: 1,
            ChangeStatus.MODIFIED: 2,
            ChangeStatus.MOVED: 3,
            ChangeStatus.RENAMED: 4,
        }
        sorted_changes = sorted(
            [c for c in comparison.ap_changes if c.status != ChangeStatus.UNCHANGED],
            key=lambda c: (status_order.get(c.status, 99), c.floor_name, c.ap_name),
        )

        for row_idx, change in enumerate(sorted_changes, 2):
            ws_changes.cell(row=row_idx, column=1, value=change.ap_name)
            status_cell = ws_changes.cell(row=row_idx, column=2, value=change.status.value)
            ws_changes.cell(row=row_idx, column=3, value=change.floor_name)
            ws_changes.cell(row=row_idx, column=4, value=change.get_details_string())

            # Apply status color
            if change.status in status_fills:
                status_cell.fill = status_fills[change.status]

        # Adjust column widths
        ws_changes.column_dimensions["A"].width = 25
        ws_changes.column_dimensions["B"].width = 12
        ws_changes.column_dimensions["C"].width = 20
        ws_changes.column_dimensions["D"].width = 50

        # Auto-filter
        ws_changes.auto_filter.ref = f"A1:D{len(sorted_changes) + 1}"

        wb.save(output_path)
        logger.info(f"Exported comparison to Excel: {output_path}")
        return output_path


class HTMLComparisonExporter(ComparisonExporter):
    """Export comparison results to HTML format."""

    def export(self, comparison: ComparisonResult, diff_images: list[Path] = None) -> Path:
        """Export comparison to HTML file.

        Args:
            comparison: ComparisonResult to export
            diff_images: Optional list of diff visualization paths to embed

        Returns:
            Path to generated HTML file
        """
        output_path = self.output_dir / f"{self.project_name}_comparison.html"

        # Embed images as base64
        embedded_images = {}
        if diff_images:
            # Handle both dict (floor_name -> path) and list formats
            img_paths = diff_images.values() if isinstance(diff_images, dict) else diff_images
            for img_path in img_paths:
                img_path = Path(img_path) if isinstance(img_path, str) else img_path
                if img_path.exists():
                    with open(img_path, "rb") as f:
                        img_data = base64.b64encode(f.read()).decode("utf-8")
                        embedded_images[img_path.stem] = img_data

        html = self._generate_html(comparison, embedded_images)

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html)

        logger.info(f"Exported comparison to HTML: {output_path}")
        return output_path

    def _generate_html(self, comparison: ComparisonResult, embedded_images: dict) -> str:
        """Generate HTML content."""
        inv = comparison.inventory_change

        # Sort changes
        status_order = {
            ChangeStatus.REMOVED: 0,
            ChangeStatus.ADDED: 1,
            ChangeStatus.MODIFIED: 2,
            ChangeStatus.MOVED: 3,
            ChangeStatus.RENAMED: 4,
        }
        sorted_changes = sorted(
            [c for c in comparison.ap_changes if c.status != ChangeStatus.UNCHANGED],
            key=lambda c: (status_order.get(c.status, 99), c.floor_name, c.ap_name),
        )

        # Generate table rows
        rows_html = ""
        for change in sorted_changes:
            status_class = change.status.value
            rows_html += f"""
            <tr class="status-{status_class}">
                <td>{change.ap_name}</td>
                <td><span class="status-badge {status_class}">{change.status.value}</span></td>
                <td>{change.floor_name}</td>
                <td>{change.get_details_string()}</td>
            </tr>
            """

        # Generate images section
        images_html = ""
        if embedded_images:
            images_html = '<div class="images-section"><h2>Floor Plan Diffs</h2>'
            for name, data in embedded_images.items():
                images_html += f"""
                <div class="floor-image">
                    <h3>{name.replace('diff_', '').replace('_', ' ')}</h3>
                    <img src="data:image/png;base64,{data}" alt="{name}">
                </div>
                """
            images_html += "</div>"

        return f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Project Comparison Report</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; padding: 20px; background: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; }}
        h1 {{ color: #333; margin-bottom: 20px; }}
        h2 {{ color: #444; margin: 20px 0 10px; }}
        .summary {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin-bottom: 30px; }}
        .card {{ background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .card h3 {{ color: #666; font-size: 14px; margin-bottom: 5px; }}
        .card .value {{ font-size: 28px; font-weight: bold; color: #333; }}
        .card.added .value {{ color: #28a745; }}
        .card.removed .value {{ color: #dc3545; }}
        .card.modified .value {{ color: #ffc107; }}
        .card.moved .value {{ color: #6f42c1; }}
        .card.renamed .value {{ color: #fd7e14; }}
        table {{ width: 100%; border-collapse: collapse; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        th {{ background: #4472C4; color: white; padding: 12px; text-align: left; }}
        td {{ padding: 10px 12px; border-bottom: 1px solid #eee; }}
        tr:hover {{ background: #f8f9fa; }}
        .status-badge {{ padding: 4px 8px; border-radius: 4px; font-size: 12px; font-weight: 500; }}
        .status-badge.added {{ background: #c6efce; color: #155724; }}
        .status-badge.removed {{ background: #ffc7ce; color: #721c24; }}
        .status-badge.modified {{ background: #ffeb9c; color: #856404; }}
        .status-badge.moved {{ background: #d9e1f2; color: #3d5a80; }}
        .status-badge.renamed {{ background: #ffd699; color: #8a5700; }}
        .images-section {{ margin-top: 30px; }}
        .floor-image {{ background: white; padding: 20px; border-radius: 8px; margin-bottom: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .floor-image img {{ max-width: 100%; height: auto; border: 1px solid #ddd; }}
        .meta {{ color: #666; font-size: 14px; margin-bottom: 20px; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>Project Comparison Report</h1>
        <p class="meta">
            <strong>{comparison.project_a_name}</strong> → <strong>{comparison.project_b_name}</strong><br>
            Generated: {comparison.comparison_timestamp.strftime("%Y-%m-%d %H:%M")}
        </p>

        <div class="summary">
            <div class="card">
                <h3>APs Before</h3>
                <div class="value">{inv.old_total_aps}</div>
            </div>
            <div class="card">
                <h3>APs After</h3>
                <div class="value">{inv.new_total_aps}</div>
            </div>
            <div class="card added">
                <h3>Added</h3>
                <div class="value">+{inv.aps_added}</div>
            </div>
            <div class="card removed">
                <h3>Removed</h3>
                <div class="value">-{inv.aps_removed}</div>
            </div>
            <div class="card modified">
                <h3>Modified</h3>
                <div class="value">{inv.aps_modified}</div>
            </div>
            <div class="card moved">
                <h3>Moved</h3>
                <div class="value">{inv.aps_moved}</div>
            </div>
            <div class="card renamed">
                <h3>Renamed</h3>
                <div class="value">{inv.aps_renamed}</div>
            </div>
        </div>

        <h2>Changes</h2>
        <table>
            <thead>
                <tr>
                    <th>AP Name</th>
                    <th>Status</th>
                    <th>Floor</th>
                    <th>Details</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>

        {images_html}
    </div>
</body>
</html>"""


class JSONComparisonExporter(ComparisonExporter):
    """Export comparison results to JSON format."""

    def export(self, comparison: ComparisonResult, pretty: bool = True) -> Path:
        """Export comparison to JSON file.

        Args:
            comparison: ComparisonResult to export
            pretty: Whether to indent JSON output

        Returns:
            Path to generated JSON file
        """
        output_path = self.output_dir / f"{self.project_name}_comparison.json"

        data = comparison.to_dict()

        with open(output_path, "w", encoding="utf-8") as f:
            if pretty:
                json.dump(data, f, indent=2, ensure_ascii=False, default=str)
            else:
                json.dump(data, f, ensure_ascii=False, default=str)

        logger.info(f"Exported comparison to JSON: {output_path}")
        return output_path


class PDFComparisonExporter(ComparisonExporter):
    """Export comparison results to PDF format."""

    def export(
        self, comparison: ComparisonResult, diff_images: list[Path] = None
    ) -> Optional[Path]:
        """Export comparison to PDF file.

        Args:
            comparison: ComparisonResult to export
            diff_images: Optional list of diff visualization paths to embed

        Returns:
            Path to generated PDF file or None if WeasyPrint unavailable
        """
        try:
            from weasyprint import HTML
        except ImportError:
            logger.warning("WeasyPrint not available, skipping PDF export")
            return None

        # Generate HTML first
        html_exporter = HTMLComparisonExporter(self.output_dir, self.project_name)
        html_content = html_exporter._generate_html(
            comparison, self._load_images_as_base64(diff_images)
        )

        output_path = self.output_dir / f"{self.project_name}_comparison.pdf"

        HTML(string=html_content).write_pdf(output_path)
        logger.info(f"Exported comparison to PDF: {output_path}")
        return output_path

    def _load_images_as_base64(self, diff_images=None) -> dict:
        """Load images as base64 for embedding."""
        embedded = {}
        if diff_images:
            # Handle both dict (floor_name -> path) and list formats
            img_paths = diff_images.values() if isinstance(diff_images, dict) else diff_images
            for img_path in img_paths:
                img_path = Path(img_path) if isinstance(img_path, str) else img_path
                if img_path.exists():
                    with open(img_path, "rb") as f:
                        img_data = base64.b64encode(f.read()).decode("utf-8")
                        embedded[img_path.stem] = img_data
        return embedded


def export_comparison(
    comparison: ComparisonResult,
    output_dir: Path,
    formats: list[str],
    project_name: str = "comparison",
    diff_images: list[Path] = None,
) -> dict[str, Path]:
    """Export comparison to multiple formats.

    Args:
        comparison: ComparisonResult to export
        output_dir: Output directory
        formats: List of formats ("csv", "excel", "html", "json", "pdf")
        project_name: Base name for output files
        diff_images: Optional list of diff visualization paths

    Returns:
        Dictionary mapping format to output path
    """
    results = {}

    for fmt in formats:
        fmt = fmt.lower()
        try:
            if fmt == "csv":
                exporter = CSVComparisonExporter(output_dir, project_name)
                results["csv"] = exporter.export(comparison)
            elif fmt == "excel":
                exporter = ExcelComparisonExporter(output_dir, project_name)
                results["excel"] = exporter.export(comparison, diff_images)
            elif fmt == "html":
                exporter = HTMLComparisonExporter(output_dir, project_name)
                results["html"] = exporter.export(comparison, diff_images)
            elif fmt == "json":
                exporter = JSONComparisonExporter(output_dir, project_name)
                results["json"] = exporter.export(comparison)
            elif fmt == "pdf":
                exporter = PDFComparisonExporter(output_dir, project_name)
                path = exporter.export(comparison, diff_images)
                if path:
                    results["pdf"] = path
            else:
                logger.warning(f"Unknown export format: {fmt}")
        except Exception as e:
            logger.error(f"Error exporting to {fmt}: {e}")

    return results
