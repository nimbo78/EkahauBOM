#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""Unit tests for Excel exporter."""

from __future__ import annotations


import pytest
from pathlib import Path
from ekahau_bom.exporters.excel_exporter import ExcelExporter
from ekahau_bom.models import ProjectData, AccessPoint, Antenna, Tag, Floor, ProjectMetadata, Radio


@pytest.fixture
def sample_project_data():
    """Create sample project data."""
    aps = [
        AccessPoint(
            id="ap1",
            vendor="Cisco",
            model="AP-515",
            color="Yellow",
            floor_name="Floor 1",
            floor_id="f1",
            mine=True,
            tags=[Tag("Location", "Building A", "1")],
        ),
        AccessPoint(
            id="ap2",
            vendor="Cisco",
            model="AP-515",
            color="Yellow",
            floor_name="Floor 1",
            floor_id="f1",
            mine=True,
            tags=[Tag("Location", "Building A", "1")],
        ),
        AccessPoint(
            id="ap3",
            vendor="Cisco",
            model="AP-635",
            color="Red",
            floor_name="Floor 2",
            floor_id="f2",
            mine=True,
            tags=[],
        ),
        AccessPoint(
            id="ap4",
            vendor="Aruba",
            model="AP-515",
            color="Yellow",
            floor_name="Floor 1",
            floor_id="f1",
            mine=True,
            tags=[],
        ),
        AccessPoint(
            id="ap5",
            vendor="Aruba",
            model="AP-635",
            color="Blue",
            floor_name="Floor 2",
            floor_id="f2",
            mine=True,
            tags=[Tag("Location", "Building B", "1")],
        ),
    ]
    antennas = [
        Antenna("ANT-2513P4M-N-R", "ant1"),
        Antenna("ANT-2513P4M-N-R", "ant1"),
        Antenna("ANT-20", "ant2"),
    ]
    floors = {"f1": Floor("f1", "Floor 1"), "f2": Floor("f2", "Floor 2")}
    return ProjectData(
        project_name="Test Project", access_points=aps, antennas=antennas, floors=floors
    )


@pytest.fixture
def detailed_project_data():
    """Create project data with detailed AP installation parameters."""
    aps = [
        AccessPoint(
            id="ap1",
            vendor="Cisco",
            model="AP-515",
            color="Yellow",
            floor_name="Floor 1",
            floor_id="f1",
            mine=True,
            name="AP-Office-01",
            location_x=10.5,
            location_y=20.3,
            mounting_height=3.0,
            azimuth=45.0,
            tilt=10.0,
            enabled=True,
            tags=[Tag("Location", "Office", "1")],
        ),
        AccessPoint(
            id="ap2",
            vendor="Aruba",
            model="AP-635",
            color="Red",
            floor_name="Floor 2",
            floor_id="f2",
            mine=True,
            name="AP-Lobby-01",
            location_x=5.2,
            location_y=15.8,
            mounting_height=2.5,
            azimuth=90.0,
            tilt=15.0,
            enabled=False,
            tags=[],
        ),
    ]
    radios = [
        Radio(
            id="radio1",
            access_point_id="ap1",
            frequency_band="5GHz",
            channel=36,
            channel_width=80,
            tx_power=20,
            antenna_direction=45.0,
            antenna_tilt=10.0,
            antenna_height=3.0,
        ),
    ]
    floors = {"f1": Floor("f1", "Floor 1"), "f2": Floor("f2", "Floor 2")}
    metadata = ProjectMetadata(
        name="Enterprise Campus Project",
        customer="Acme Corporation",
        location="Building A, Floor 3",
        responsible_person="John Doe",
        schema_version="1.0",
    )
    return ProjectData(
        project_name="Detailed Project",
        access_points=aps,
        antennas=[],
        floors=floors,
        radios=radios,
        metadata=metadata,
    )


class TestExcelExporter:
    """Test ExcelExporter class."""

    def test_export_creates_file(self, sample_project_data, tmp_path):
        """Test that export creates Excel file."""
        exporter = ExcelExporter(tmp_path)
        files = exporter.export(sample_project_data)

        assert len(files) == 1
        assert files[0].exists()
        assert files[0].suffix == ".xlsx"
        assert "Test Project" in files[0].name

    def test_export_has_required_sheets(self, sample_project_data, tmp_path):
        """Test that Excel file has all required sheets."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        exporter = ExcelExporter(tmp_path)
        files = exporter.export(sample_project_data)

        wb = load_workbook(files[0])
        sheet_names = wb.sheetnames

        assert "Summary" in sheet_names
        assert "Access Points" in sheet_names
        assert "Antennas" in sheet_names
        assert "By Floor" in sheet_names
        assert "By Color" in sheet_names
        assert "By Vendor" in sheet_names
        assert "By Model" in sheet_names

    def test_summary_sheet_content(self, sample_project_data, tmp_path):
        """Test Summary sheet has correct data."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        exporter = ExcelExporter(tmp_path)
        files = exporter.export(sample_project_data)

        wb = load_workbook(files[0])
        ws = wb["Summary"]

        # Check title
        assert ws["A1"].value == "Project Summary"

        # Find where "Project Statistics" section starts (after optional metadata)
        stats_row = None
        for row in range(1, 20):
            if ws[f"A{row}"].value == "Project Statistics":
                stats_row = row
                break

        assert stats_row is not None, "Project Statistics section not found"

        # Check counts (relative to stats_row)
        file_name_row = stats_row + 1
        ap_row = stats_row + 2
        antenna_row = stats_row + 3

        assert ws[f"A{file_name_row}"].value == "File Name:"
        assert ws[f"B{file_name_row}"].value == "Test Project"

        assert ws[f"A{ap_row}"].value == "Total Access Points:"
        assert ws[f"B{ap_row}"].value == 5

        assert ws[f"A{antenna_row}"].value == "Total Antennas:"
        assert ws[f"B{antenna_row}"].value == 3

    def test_access_points_sheet_content(self, sample_project_data, tmp_path):
        """Test Access Points sheet has correct data."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        exporter = ExcelExporter(tmp_path)
        files = exporter.export(sample_project_data)

        wb = load_workbook(files[0])
        ws = wb["Access Points"]

        # Check headers
        headers = [cell.value for cell in ws[1]]
        assert "Vendor" in headers
        assert "Model" in headers
        assert "Floor" in headers
        assert "Color" in headers
        assert "Tags" in headers
        assert "Quantity" in headers

        # Check we have data rows (header + data)
        assert ws.max_row > 1

    def test_antennas_sheet_content(self, sample_project_data, tmp_path):
        """Test Antennas sheet has correct data."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        exporter = ExcelExporter(tmp_path)
        files = exporter.export(sample_project_data)

        wb = load_workbook(files[0])
        ws = wb["Antennas"]

        # Check headers
        headers = [cell.value for cell in ws[1]]
        assert "Antenna Model" in headers
        assert "Quantity" in headers

        # Check we have data rows
        assert ws.max_row > 1

    def test_grouped_sheets_have_charts(self, sample_project_data, tmp_path):
        """Test grouped sheets have charts."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        exporter = ExcelExporter(tmp_path)
        files = exporter.export(sample_project_data)

        wb = load_workbook(files[0])

        # Check that grouped sheets have charts
        for sheet_name in ["By Floor", "By Color", "By Vendor", "By Model"]:
            ws = wb[sheet_name]
            # openpyxl stores charts in _charts attribute
            assert len(ws._charts) > 0, f"{sheet_name} should have at least one chart"

    def test_format_name(self):
        """Test format name property."""
        exporter = ExcelExporter(Path("."))
        assert exporter.format_name == "Excel"

    def test_empty_project_data(self, tmp_path):
        """Test export with empty project data."""
        empty_data = ProjectData(
            project_name="Empty Project", access_points=[], antennas=[], floors={}
        )
        exporter = ExcelExporter(tmp_path)
        files = exporter.export(empty_data)

        assert len(files) == 1
        assert files[0].exists()

    def test_tags_in_access_points_export(self, sample_project_data, tmp_path):
        """Test that tags are correctly exported in Access Points sheet."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        exporter = ExcelExporter(tmp_path)
        files = exporter.export(sample_project_data)

        wb = load_workbook(files[0])
        ws = wb["Access Points"]

        # Find Tags column
        headers = [cell.value for cell in ws[1]]
        tags_col_idx = headers.index("Tags") + 1

        # Check that some rows have tags
        has_tags = False
        for row_idx in range(2, ws.max_row + 1):
            tags_cell = ws.cell(row=row_idx, column=tags_col_idx)
            if tags_cell.value and "Location" in tags_cell.value:
                has_tags = True
                # Verify format is "Key:Value"
                assert ":" in tags_cell.value
                break

        assert has_tags, "At least one AP should have tags exported"

    def test_detailed_access_points_sheet(self, detailed_project_data, tmp_path):
        """Test AP Installation Details sheet with coordinates and mounting info."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        exporter = ExcelExporter(tmp_path)
        files = exporter.export(detailed_project_data)

        wb = load_workbook(files[0])

        # Check sheet exists
        assert "AP Installation Details" in wb.sheetnames

        ws = wb["AP Installation Details"]

        # Check headers
        headers = [cell.value for cell in ws[1]]
        assert "AP Name" in headers
        assert "Vendor" in headers
        assert "Model" in headers
        assert "Floor" in headers
        assert "Location X (m)" in headers
        assert "Location Y (m)" in headers
        assert "Mounting Height (m)" in headers
        assert "Azimuth (°)" in headers
        assert "Tilt (°)" in headers
        assert "Color" in headers
        assert "Tags" in headers
        assert "Enabled" in headers

        # Check data rows (should have 2 APs)
        assert ws.max_row == 3  # header + 2 APs

        # Verify AP-Office-01 data
        assert ws.cell(2, headers.index("AP Name") + 1).value == "AP-Office-01"
        assert ws.cell(2, headers.index("Location X (m)") + 1).value == 10.5
        assert ws.cell(2, headers.index("Location Y (m)") + 1).value == 20.3
        assert ws.cell(2, headers.index("Mounting Height (m)") + 1).value == 3.0
        assert ws.cell(2, headers.index("Azimuth (°)") + 1).value == 45.0
        assert ws.cell(2, headers.index("Tilt (°)") + 1).value == 10.0
        assert ws.cell(2, headers.index("Enabled") + 1).value == "Yes"

        # Verify AP-Lobby-01 data
        assert ws.cell(3, headers.index("AP Name") + 1).value == "AP-Lobby-01"
        assert ws.cell(3, headers.index("Enabled") + 1).value == "No"

    def test_analytics_sheet_with_radios(self, detailed_project_data, tmp_path):
        """Test Analytics sheet creation when radio data is available."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        exporter = ExcelExporter(tmp_path)
        files = exporter.export(detailed_project_data)

        wb = load_workbook(files[0])

        # Check Analytics sheet exists (should be created because we have radios)
        assert "Analytics" in wb.sheetnames

        ws = wb["Analytics"]

        # Check for analytics sections
        # Look for key section headers
        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend([str(cell) for cell in row if cell is not None])

        content = " ".join(all_values)

        # Should have some analytics content
        assert len(content) > 0
        assert ws.max_row > 1  # Should have more than just headers

    def test_summary_sheet_with_metadata(self, detailed_project_data, tmp_path):
        """Test Summary sheet includes project metadata."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        exporter = ExcelExporter(tmp_path)
        files = exporter.export(detailed_project_data)

        wb = load_workbook(files[0])
        ws = wb["Summary"]

        # Collect all cell values
        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend([str(cell) for cell in row if cell is not None])

        content = " ".join(all_values)

        # Check for metadata content
        assert "Enterprise Campus Project" in content
        assert "Acme Corporation" in content
        assert "Building A, Floor 3" in content
        assert "John Doe" in content
        assert "Project Information" in content

    def test_export_without_metadata(self, sample_project_data, tmp_path):
        """Test export works correctly without metadata."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        exporter = ExcelExporter(tmp_path)
        files = exporter.export(sample_project_data)

        wb = load_workbook(files[0])
        ws = wb["Summary"]

        # Should still have Project Statistics section
        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend([str(cell) for cell in row if cell is not None])

        content = " ".join(all_values)
        assert "Project Statistics" in content
        assert "Total Access Points:" in content

    def test_export_with_radios_creates_analytics_sheet(self, detailed_project_data, tmp_path):
        """Test that Analytics sheet is created when radios data is available."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        exporter = ExcelExporter(tmp_path)
        files = exporter.export(detailed_project_data)

        wb = load_workbook(files[0])
        sheet_names = wb.sheetnames

        # Analytics sheet should exist because we have radios
        assert "Analytics" in sheet_names

    def test_export_without_radios_no_analytics_sheet(self, sample_project_data, tmp_path):
        """Test that Analytics sheet is NOT created when no radios/mounting data."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        exporter = ExcelExporter(tmp_path)
        files = exporter.export(sample_project_data)

        wb = load_workbook(files[0])
        sheet_names = wb.sheetnames

        # Analytics sheet should NOT exist (no radios, no mounting height)
        assert "Analytics" not in sheet_names

    def test_number_formatting_in_detailed_sheet(self, detailed_project_data, tmp_path):
        """Test that numeric columns have proper number formatting."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")

        exporter = ExcelExporter(tmp_path)
        files = exporter.export(detailed_project_data)

        wb = load_workbook(files[0])
        ws = wb["AP Installation Details"]

        headers = [cell.value for cell in ws[1]]

        # Check number format for Location X (should be 0.00)
        loc_x_col = headers.index("Location X (m)") + 1
        assert ws.cell(2, loc_x_col).number_format == "0.00"

        # Check number format for Mounting Height (should be 0.00)
        height_col = headers.index("Mounting Height (m)") + 1
        assert ws.cell(2, height_col).number_format == "0.00"

        # Check number format for Azimuth (should be 0.0)
        azimuth_col = headers.index("Azimuth (°)") + 1
        assert ws.cell(2, azimuth_col).number_format == "0.0"

    def test_export_with_aps_requiring_height_adjustment(self, tmp_path):
        """Test analytics sheet with APs requiring height adjustment."""
        from ekahau_bom.models import Radio, Floor
        from openpyxl import load_workbook

        # Create APs with different mounting heights
        aps = [
            AccessPoint(
                id="ap1",
                vendor="Cisco",
                model="AP-515",
                color="Yellow",
                floor_name="Floor 1",
                floor_id="f1",
                mine=True,
                mounting_height=2.0,
                azimuth=0.0,
                tilt=0.0,
            ),
            AccessPoint(
                id="ap2",
                vendor="Cisco",
                model="AP-515",
                color="Yellow",
                floor_name="Floor 1",
                floor_id="f1",
                mine=True,
                mounting_height=5.0,
                azimuth=0.0,
                tilt=0.0,  # Different height
            ),
        ]

        floors = {"f1": Floor(id="f1", name="Floor 1")}

        radios = [
            Radio(id="r1", access_point_id="ap1", antenna_height=2.0),
            Radio(id="r2", access_point_id="ap2", antenna_height=5.0),
        ]

        project_data = ProjectData(
            project_name="Height Adjustment Test",
            access_points=aps,
            antennas=[],
            floors=floors,
            radios=radios,
        )

        exporter = ExcelExporter(tmp_path)
        files = exporter.export(project_data)

        wb = load_workbook(files[0])
        assert "Analytics" in wb.sheetnames

        ws = wb["Analytics"]

        # Find the row with "APs Requiring Height Adjustment"
        found = False
        for row in ws.iter_rows():
            if row[0].value == "APs Requiring Height Adjustment":
                found = True
                # Should have >= 0 APs requiring adjustment
                assert row[1].value >= 0
                # Check if cell is highlighted when value > 0
                if row[1].value > 0:
                    assert row[1].fill.start_color.rgb is not None
                break

        assert found, "Should have 'APs Requiring Height Adjustment' row"

    def test_export_with_save_error(self, tmp_path):
        """Test error handling when Excel file cannot be saved."""
        from unittest.mock import patch, MagicMock

        aps = [
            AccessPoint(
                vendor="Cisco", model="AP-515", color="Yellow", floor_name="Floor 1", tags=[]
            )
        ]
        project_data = ProjectData(
            project_name="Error Test", access_points=aps, antennas=[], floors={}
        )

        exporter = ExcelExporter(tmp_path)

        # Mock workbook.save to raise an exception
        with patch("openpyxl.Workbook.save", side_effect=PermissionError("File is locked")):
            with pytest.raises(IOError, match="Failed to save Excel file"):
                exporter.export(project_data)

    def test_auto_size_columns_with_exception(self, tmp_path):
        """Test _auto_size_columns handles exceptions in cell values."""
        from openpyxl import Workbook

        aps = [
            AccessPoint(
                vendor="Cisco", model="AP-515", color="Yellow", floor_name="Floor 1", tags=[]
            )
        ]
        project_data = ProjectData(
            project_name="AutoSize Test", access_points=aps, antennas=[], floors={}
        )

        exporter = ExcelExporter(tmp_path)

        # Create a workbook with a problematic cell
        wb = Workbook()
        ws = wb.active

        # Add normal cells
        ws["A1"] = "Normal Text"
        ws["B1"] = 12345

        # Create a cell with a value that might cause issues when converting to string
        # This should be handled by the except block
        ws["C1"] = None

        # Call _auto_size_columns - should not raise exception
        exporter._auto_size_columns(ws)

        # Verify columns were sized (should complete without errors)
        assert ws.column_dimensions["A"].width > 0

    def test_analytics_with_wifi_standards_section(self, tmp_path):
        """Test that Wi-Fi Standards section is generated in analytics."""
        from ekahau_bom.models import Radio, Floor
        from openpyxl import load_workbook

        # Create APs with radios that have Wi-Fi standards
        aps = [
            AccessPoint(
                id=f"ap{i}",
                vendor="Cisco",
                model="AP-515",
                color="Yellow",
                floor_name="Floor 1",
                floor_id="f1",
                mine=True,
                mounting_height=3.0,
                azimuth=0.0,
                tilt=0.0,  # Add mounting info
            )
            for i in range(1, 6)
        ]

        floors = {"f1": Floor(id="f1", name="Floor 1")}

        # Create radios with different Wi-Fi standards
        radios = [
            Radio(
                id="r1",
                access_point_id="ap1",
                frequency_band="5GHz",
                channel=36,
                channel_width=80,
                tx_power=20.0,
                antenna_height=3.0,
                standard="802.11ax",
            ),
            Radio(
                id="r2",
                access_point_id="ap2",
                frequency_band="5GHz",
                channel=40,
                channel_width=80,
                tx_power=20.0,
                antenna_height=3.0,
                standard="802.11ac",
            ),
            Radio(
                id="r3",
                access_point_id="ap3",
                frequency_band="5GHz",
                channel=44,
                channel_width=80,
                tx_power=20.0,
                antenna_height=3.0,
                standard="802.11ax",
            ),
            Radio(
                id="r4",
                access_point_id="ap4",
                frequency_band="2.4GHz",
                channel=6,
                channel_width=20,
                tx_power=17.0,
                antenna_height=3.0,
                standard="802.11n",
            ),
            Radio(
                id="r5",
                access_point_id="ap5",
                frequency_band="5GHz",
                channel=48,
                channel_width=80,
                tx_power=20.0,
                antenna_height=3.0,
                standard="802.11ac",
            ),
        ]

        project_data = ProjectData(
            project_name="WiFi Standards Test",
            access_points=aps,
            antennas=[],
            floors=floors,
            radios=radios,
        )

        exporter = ExcelExporter(tmp_path)
        files = exporter.export(project_data)

        wb = load_workbook(files[0])
        ws = wb["Analytics"]

        # Search for Wi-Fi Standards section
        found_standards = False
        found_802_11ax = False
        for row in ws.iter_rows():
            if row[0].value == "Wi-Fi Standards":
                found_standards = True
            if row[0].value == "802.11ax":
                found_802_11ax = True

        assert found_standards, "Should have 'Wi-Fi Standards' section"
        assert found_802_11ax, "Should have '802.11ax' in standards list"

    def test_auto_size_columns_with_actual_exception(self, tmp_path):
        """Test that _auto_size_columns except block is covered."""
        from openpyxl import Workbook
        from unittest.mock import PropertyMock, patch

        exporter = ExcelExporter(tmp_path)
        wb = Workbook()
        ws = wb.active

        # Add normal cells
        ws["A1"] = "Normal"
        ws["B1"] = "Test"
        ws["C1"] = "Data"

        # Get the actual cell object
        problem_cell = ws["C1"]

        # Create a mock that raises exception when value property is accessed
        with patch.object(type(problem_cell), "value", new_callable=PropertyMock) as mock_value:
            mock_value.side_effect = RuntimeError("Simulated cell value error")

            # This should not raise, exception should be caught by except block
            try:
                exporter._auto_size_columns(ws)
            except RuntimeError:
                # If exception propagated, the except block didn't work
                pytest.fail("Exception was not caught by except block")

        # Verify it completed successfully
        assert ws.column_dimensions["A"].width > 0
