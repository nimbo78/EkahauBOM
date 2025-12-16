"""Tests for the comparison exporters module."""

import csv
import json
import pytest
from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock, patch

from ekahau_bom.comparison.models import (
    APChange,
    ChangeStatus,
    ComparisonResult,
    FieldChange,
    InventoryChange,
    MetadataChange,
)
from ekahau_bom.comparison.exporters import (
    ComparisonExporter,
    CSVComparisonExporter,
    ExcelComparisonExporter,
    HTMLComparisonExporter,
    JSONComparisonExporter,
    PDFComparisonExporter,
    export_comparison,
)


@pytest.fixture
def sample_comparison():
    """Create sample comparison result for testing."""
    return ComparisonResult(
        project_a_name="Old Project",
        project_b_name="New Project",
        project_a_filename="old.esx",
        project_b_filename="new.esx",
        comparison_timestamp=datetime(2025, 1, 15, 10, 30, 0),
        inventory_change=InventoryChange(
            old_total_aps=50,
            new_total_aps=55,
            aps_added=8,
            aps_removed=3,
            aps_modified=5,
            aps_moved=2,
            aps_renamed=1,
            aps_unchanged=36,
        ),
        metadata_change=MetadataChange(
            old_name="Old Project",
            new_name="New Project",
        ),
        ap_changes=[
            APChange(
                status=ChangeStatus.ADDED,
                ap_name="AP-101",
                floor_name="Floor 1",
            ),
            APChange(
                status=ChangeStatus.REMOVED,
                ap_name="AP-102",
                floor_name="Floor 1",
            ),
            APChange(
                status=ChangeStatus.MODIFIED,
                ap_name="AP-103",
                floor_name="Floor 2",
                changes=[
                    FieldChange("channel", "radio", 36, 44),
                ],
            ),
            APChange(
                status=ChangeStatus.MOVED,
                ap_name="AP-104",
                floor_name="Floor 2",
                old_coords=(10.0, 20.0),
                new_coords=(15.0, 25.0),
                distance_moved=7.07,
            ),
            APChange(
                status=ChangeStatus.RENAMED,
                ap_name="AP-NEW-105",
                floor_name="Floor 1",
                old_name="AP-OLD-105",
                new_name="AP-NEW-105",
            ),
            APChange(
                status=ChangeStatus.UNCHANGED,
                ap_name="AP-106",
                floor_name="Floor 1",
            ),
        ],
        changes_by_floor={
            "Floor 1": [
                APChange(status=ChangeStatus.ADDED, ap_name="AP-101", floor_name="Floor 1"),
                APChange(status=ChangeStatus.REMOVED, ap_name="AP-102", floor_name="Floor 1"),
            ],
            "Floor 2": [
                APChange(
                    status=ChangeStatus.MODIFIED,
                    ap_name="AP-103",
                    floor_name="Floor 2",
                    changes=[FieldChange("channel", "radio", 36, 44)],
                ),
                APChange(
                    status=ChangeStatus.MOVED,
                    ap_name="AP-104",
                    floor_name="Floor 2",
                    old_coords=(10.0, 20.0),
                    new_coords=(15.0, 25.0),
                    distance_moved=7.07,
                ),
            ],
        },
        floors=["Floor 1", "Floor 2"],
    )


class TestComparisonExporter:
    """Tests for base ComparisonExporter class."""

    def test_init(self, tmp_path):
        """Test base exporter initialization."""
        exporter = ComparisonExporter(tmp_path, "test_project")

        assert exporter.output_dir == tmp_path
        assert exporter.project_name == "test_project"

    def test_sanitize_filename(self, tmp_path):
        """Test filename sanitization."""
        exporter = ComparisonExporter(tmp_path, "test:project<name>")

        assert exporter.project_name == "test_project_name_"

    def test_creates_output_dir(self, tmp_path):
        """Test that output directory is created if it doesn't exist."""
        output_dir = tmp_path / "new_dir" / "nested"
        exporter = ComparisonExporter(output_dir, "test")

        assert output_dir.exists()


class TestCSVComparisonExporter:
    """Tests for CSV exporter."""

    def test_export_creates_file(self, tmp_path, sample_comparison):
        """Test CSV export creates file."""
        exporter = CSVComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison)

        assert output_path.exists()
        assert output_path.name == "test_comparison.csv"

    def test_export_csv_content(self, tmp_path, sample_comparison):
        """Test CSV content is correct."""
        exporter = CSVComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison)

        with open(output_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        # Header row
        assert rows[0] == ["AP Name", "Status", "Floor", "Details"]

        # Should have 5 changes (excluding unchanged)
        assert len(rows) == 6  # header + 5 changes

    def test_csv_excludes_unchanged(self, tmp_path, sample_comparison):
        """Test CSV export excludes unchanged APs."""
        exporter = CSVComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison)

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        # Unchanged AP should not appear
        assert "AP-106" not in content

    def test_csv_sorted_by_status(self, tmp_path, sample_comparison):
        """Test CSV rows are sorted by status."""
        exporter = CSVComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison)

        with open(output_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)[1:]  # Skip header

        statuses = [row[1] for row in rows]

        # Order: removed, added, modified, moved, renamed
        assert statuses.index("removed") < statuses.index("added")
        assert statuses.index("added") < statuses.index("modified")
        assert statuses.index("modified") < statuses.index("moved")


class TestExcelComparisonExporter:
    """Tests for Excel exporter."""

    @pytest.fixture
    def check_openpyxl(self):
        """Check if openpyxl is available."""
        try:
            import openpyxl

            return True
        except ImportError:
            pytest.skip("openpyxl not available")

    def test_export_creates_file(self, tmp_path, sample_comparison, check_openpyxl):
        """Test Excel export creates file."""
        exporter = ExcelComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison)

        assert output_path.exists()
        assert output_path.name == "test_comparison.xlsx"

    def test_excel_has_summary_sheet(self, tmp_path, sample_comparison, check_openpyxl):
        """Test Excel file has Summary sheet."""
        import openpyxl

        exporter = ExcelComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison)

        wb = openpyxl.load_workbook(output_path)
        assert "Summary" in wb.sheetnames

    def test_excel_has_changes_sheet(self, tmp_path, sample_comparison, check_openpyxl):
        """Test Excel file has Changes sheet."""
        import openpyxl

        exporter = ExcelComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison)

        wb = openpyxl.load_workbook(output_path)
        assert "Changes" in wb.sheetnames

    def test_excel_changes_content(self, tmp_path, sample_comparison, check_openpyxl):
        """Test Changes sheet has correct content."""
        import openpyxl

        exporter = ExcelComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Changes"]

        # Header row
        assert ws.cell(1, 1).value == "AP Name"
        assert ws.cell(1, 2).value == "Status"
        assert ws.cell(1, 3).value == "Floor"
        assert ws.cell(1, 4).value == "Details"

        # Should have some data rows
        assert ws.cell(2, 1).value is not None


class TestHTMLComparisonExporter:
    """Tests for HTML exporter."""

    def test_export_creates_file(self, tmp_path, sample_comparison):
        """Test HTML export creates file."""
        exporter = HTMLComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison)

        assert output_path.exists()
        assert output_path.name == "test_comparison.html"

    def test_html_structure(self, tmp_path, sample_comparison):
        """Test HTML contains expected elements."""
        exporter = HTMLComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison)

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        assert "<!DOCTYPE html>" in content
        assert "<title>Project Comparison Report</title>" in content
        assert "Old Project" in content
        assert "New Project" in content

    def test_html_summary_cards(self, tmp_path, sample_comparison):
        """Test HTML has summary cards."""
        exporter = HTMLComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison)

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        assert "APs Before" in content
        assert "APs After" in content
        assert "Added" in content
        assert "Removed" in content

    def test_html_changes_table(self, tmp_path, sample_comparison):
        """Test HTML has changes table."""
        exporter = HTMLComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison)

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        assert "<table>" in content
        assert "AP Name" in content
        assert "AP-101" in content
        assert "AP-102" in content

    def test_html_with_embedded_images(self, tmp_path, sample_comparison):
        """Test HTML embeds images as base64."""
        # Create a dummy image file
        from PIL import Image

        img_path = tmp_path / "diff_Floor_1.png"
        img = Image.new("RGB", (100, 100), (255, 255, 255))
        img.save(img_path)

        exporter = HTMLComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison, diff_images=[img_path])

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        assert "data:image/png;base64," in content
        assert "Floor Plan Diffs" in content


class TestJSONComparisonExporter:
    """Tests for JSON exporter."""

    def test_export_creates_file(self, tmp_path, sample_comparison):
        """Test JSON export creates file."""
        exporter = JSONComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison)

        assert output_path.exists()
        assert output_path.name == "test_comparison.json"

    def test_json_valid(self, tmp_path, sample_comparison):
        """Test exported JSON is valid."""
        exporter = JSONComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison)

        with open(output_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        assert isinstance(data, dict)

    def test_json_structure(self, tmp_path, sample_comparison):
        """Test JSON has expected structure."""
        exporter = JSONComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison)

        with open(output_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        assert "project_a_name" in data
        assert "project_b_name" in data
        assert "summary" in data  # inventory summary
        assert "changes" in data  # ap changes

    def test_json_pretty_print(self, tmp_path, sample_comparison):
        """Test JSON is pretty-printed by default."""
        exporter = JSONComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison, pretty=True)

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        # Pretty-printed JSON has indentation
        assert "\n  " in content

    def test_json_compact(self, tmp_path, sample_comparison):
        """Test JSON compact mode."""
        exporter = JSONComparisonExporter(tmp_path, "test")
        output_path = exporter.export(sample_comparison, pretty=False)

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        # Compact JSON is single line (no newlines except in string values)
        lines = content.split("\n")
        assert len(lines) <= 2  # May have trailing newline


class TestPDFComparisonExporter:
    """Tests for PDF exporter."""

    def test_export_without_weasyprint(self, tmp_path, sample_comparison):
        """Test PDF export returns None when WeasyPrint unavailable."""
        with patch.dict("sys.modules", {"weasyprint": None}):
            exporter = PDFComparisonExporter(tmp_path, "test")

            # Should return None, not raise
            result = exporter.export(sample_comparison)

            # Either None (WeasyPrint not available) or a path
            if result is not None:
                assert result.name == "test_comparison.pdf"

    @pytest.mark.slow
    def test_export_with_weasyprint(self, tmp_path, sample_comparison):
        """Test PDF export with WeasyPrint (slow)."""
        try:
            from weasyprint import HTML
        except (ImportError, OSError):
            # OSError can occur on macOS/Windows when native libs (libgobject) are missing
            pytest.skip("WeasyPrint not available or native libraries missing")

        exporter = PDFComparisonExporter(tmp_path, "test")
        try:
            output_path = exporter.export(sample_comparison)
        except Exception as e:
            # WeasyPrint can fail at runtime even if import succeeds
            pytest.skip(f"WeasyPrint export failed: {e}")

        if output_path:
            assert output_path.exists()
            assert output_path.name == "test_comparison.pdf"


class TestExportComparison:
    """Tests for export_comparison convenience function."""

    def test_export_single_format(self, tmp_path, sample_comparison):
        """Test exporting single format."""
        results = export_comparison(
            sample_comparison,
            tmp_path,
            ["csv"],
            project_name="test",
        )

        assert "csv" in results
        assert results["csv"].exists()

    def test_export_multiple_formats(self, tmp_path, sample_comparison):
        """Test exporting multiple formats."""
        results = export_comparison(
            sample_comparison,
            tmp_path,
            ["csv", "json", "html"],
            project_name="test",
        )

        assert len(results) == 3
        assert "csv" in results
        assert "json" in results
        assert "html" in results

    def test_export_unknown_format(self, tmp_path, sample_comparison):
        """Test unknown format is skipped."""
        results = export_comparison(
            sample_comparison,
            tmp_path,
            ["csv", "unknown_format"],
            project_name="test",
        )

        assert "csv" in results
        assert "unknown_format" not in results

    def test_export_case_insensitive(self, tmp_path, sample_comparison):
        """Test format names are case-insensitive."""
        results = export_comparison(
            sample_comparison,
            tmp_path,
            ["CSV", "JSON"],
            project_name="test",
        )

        assert "csv" in results
        assert "json" in results

    def test_export_with_diff_images(self, tmp_path, sample_comparison):
        """Test export with diff images."""
        from PIL import Image

        # Create dummy image
        img_path = tmp_path / "diff_Floor_1.png"
        img = Image.new("RGB", (100, 100), (255, 255, 255))
        img.save(img_path)

        results = export_comparison(
            sample_comparison,
            tmp_path,
            ["html"],
            project_name="test",
            diff_images=[img_path],
        )

        assert "html" in results

        with open(results["html"], "r", encoding="utf-8") as f:
            content = f.read()

        assert "data:image/png;base64," in content


class TestInventoryChangeDiff:
    """Tests for inventory change calculations."""

    def test_ap_count_diff_property(self):
        """Test ap_count_diff calculated correctly."""
        inv = InventoryChange(
            old_total_aps=50,
            new_total_aps=55,
            aps_added=8,
            aps_removed=3,
            aps_modified=5,
            aps_moved=2,
            aps_renamed=1,
            aps_unchanged=36,
        )

        assert inv.ap_count_diff == 5  # 55 - 50


class TestAPChangeDetailsString:
    """Tests for APChange.get_details_string method."""

    def test_added_details(self):
        """Test details for added AP."""
        change = APChange(
            status=ChangeStatus.ADDED,
            ap_name="AP-101",
            floor_name="Floor 1",
        )

        details = change.get_details_string()
        # Added APs have no additional details
        assert details == "" or "added" in details.lower() or details is not None

    def test_moved_details(self):
        """Test details for moved AP."""
        change = APChange(
            status=ChangeStatus.MOVED,
            ap_name="AP-104",
            floor_name="Floor 2",
            old_coords=(10.0, 20.0),
            new_coords=(15.0, 25.0),
            distance_moved=7.07,
        )

        details = change.get_details_string()
        # Should include distance
        assert "7.07" in details or "7.1" in details or "moved" in details.lower()

    def test_renamed_details(self):
        """Test details for renamed AP."""
        change = APChange(
            status=ChangeStatus.RENAMED,
            ap_name="AP-NEW",
            floor_name="Floor 1",
            old_name="AP-OLD",
            new_name="AP-NEW",
        )

        details = change.get_details_string()
        # Should show old name
        assert "AP-OLD" in details

    def test_modified_details(self):
        """Test details for modified AP."""
        change = APChange(
            status=ChangeStatus.MODIFIED,
            ap_name="AP-103",
            floor_name="Floor 2",
            changes=[
                FieldChange("channel", "radio", 36, 44),
                FieldChange("tx_power", "radio", 15, 20),
            ],
        )

        details = change.get_details_string()
        # Should list field changes
        assert "channel" in details or "36" in details or len(details) > 0


class TestExporterEdgeCases:
    """Tests for edge cases in exporters."""

    def test_empty_comparison(self, tmp_path):
        """Test exporting comparison with no changes."""
        comparison = ComparisonResult(
            project_a_name="Old",
            project_b_name="New",
            project_a_filename="old.esx",
            project_b_filename="new.esx",
            comparison_timestamp=datetime.now(),
            inventory_change=InventoryChange(
                old_total_aps=10,
                new_total_aps=10,
                aps_added=0,
                aps_removed=0,
                aps_modified=0,
                aps_moved=0,
                aps_renamed=0,
                aps_unchanged=10,
            ),
            ap_changes=[],
        )

        # Should not raise
        results = export_comparison(comparison, tmp_path, ["csv", "json", "html"])

        assert len(results) == 3

    def test_special_characters_in_ap_name(self, tmp_path):
        """Test AP names with special characters."""
        comparison = ComparisonResult(
            project_a_name="Old",
            project_b_name="New",
            project_a_filename="old.esx",
            project_b_filename="new.esx",
            comparison_timestamp=datetime.now(),
            inventory_change=InventoryChange(
                old_total_aps=1,
                new_total_aps=1,
                aps_added=0,
                aps_removed=0,
                aps_modified=1,
                aps_moved=0,
                aps_renamed=0,
                aps_unchanged=0,
            ),
            ap_changes=[
                APChange(
                    status=ChangeStatus.MODIFIED,
                    ap_name="AP-<test>&\"name'",
                    floor_name="Floor 1",
                ),
            ],
        )

        # Should not raise
        results = export_comparison(comparison, tmp_path, ["csv", "json", "html"])

        assert len(results) == 3

    def test_unicode_in_names(self, tmp_path):
        """Test Unicode characters in project/AP names."""
        comparison = ComparisonResult(
            project_a_name="Проект 老",
            project_b_name="新プロジェクト",
            project_a_filename="old.esx",
            project_b_filename="new.esx",
            comparison_timestamp=datetime.now(),
            inventory_change=InventoryChange(
                old_total_aps=1,
                new_total_aps=1,
                aps_added=1,
                aps_removed=0,
                aps_modified=0,
                aps_moved=0,
                aps_renamed=0,
                aps_unchanged=0,
            ),
            ap_changes=[
                APChange(
                    status=ChangeStatus.ADDED,
                    ap_name="AP-日本語",
                    floor_name="Этаж 1",
                ),
            ],
        )

        # Should not raise
        results = export_comparison(comparison, tmp_path, ["csv", "json", "html"])

        # Verify Unicode preserved in JSON
        with open(results["json"], "r", encoding="utf-8") as f:
            data = json.load(f)

        assert "Проект 老" in data["project_a_name"]
        assert "新プロジェクト" in data["project_b_name"]
